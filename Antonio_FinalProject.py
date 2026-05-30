import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op

def clear_entries():
    cname_entry.delete(0, tk.END)
    item_prod.set("")
    qty_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)

def input_validation():
    name = cname_entry.get()
    product = item_prod.get()
    qty = qty_entry.get()
    price = price_entry.get()

    if not name or not product or not qty or not price:
        messagebox.showerror("Error", "Please fill in all fields.")
        return False
    
    if not qty.isdigit() or not price.isdigit():
        messagebox.showerror("Error", "Quantity and Price must be numbers.")
        return False

    return True

def display_order():
    wbk = op.load_workbook("Antonio_Database.xlsx")
    sheet = wbk.active

    for item in table.get_children():
        table.delete(item)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", "end", values=row)

def append_excel():
    if not input_validation():
        return
    
    name = cname_entry.get()
    product = item_prod.get()
    qty = int(qty_entry.get())
    price = int(price_entry.get())
    total_amount = qty * price
    
    wbk = op.load_workbook("Antonio_Database.xlsx")
    sheet = wbk.active

    if sheet.max_row == 1:
        new_order_id = 1
    else:
        new_order_id = sheet.cell(row=sheet.max_row, column=1).value + 1

    sheet.append([new_order_id, name, product, qty, price, total_amount])

    messagebox.showinfo("Success", "Order added successfully!")

    wbk.save("Antonio_Database.xlsx")
    display_order()
    clear_entries()

def auto_populate(event):
    selected = table.focus()
    values = table.item(selected, "values")

    if values:
        cname_entry.delete(0, tk.END)
        item_prod.delete(0, tk.END)
        qty_entry.delete(0, tk.END)
        price_entry.delete(0, tk.END)

        cname_entry.insert(0, values[1])
        item_prod.insert(0, values[2])
        qty_entry.insert(0, values[3])
        price_entry.insert(0, values[4])

def update():
    selected = table.focus()
    
    if not input_validation():
        return

    if not selected:
        messagebox.showerror("Error", "Please select an order to update.")
        return

    values = table.item(selected, "values")
    record_order_id = values[0]

    wbk = op.load_workbook("Antonio_Database.xlsx")
    sheet = wbk.active

    if values:
    
        for row in sheet.iter_rows(min_row=2):

            if str(row[0].value) == str(record_order_id):
                row[1].value = cname_entry.get()
                row[2].value = item_prod.get()
                row[3].value = int(qty_entry.get())
                row[4].value = int(price_entry.get())
                row[5].value = int(qty_entry.get()) * int(price_entry.get())

        wbk.save("Antonio_Database.xlsx")

        display_order()

    messagebox.showinfo("Success", "Order updated successfully!")

    clear_entries()

def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Please select a record first.")
        return
    
    values = table.item(selected, "values")
    record_id = values[0]

    confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this record?")
    if not confirm:
        return

    wbk = op.load_workbook("Antonio_Database.xlsx")
    sheet = wbk.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(record_id):
            sheet.delete_rows(i)
            break
            
    wbk.save("Antonio_Database.xlsx")

    messagebox.showinfo("Success", "Record deleted successfully!")

    display_order()

    clear_entries()

def auto_price(event):
    if item_prod.get() == "Spanish Latte":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "120")
    elif item_prod.get() == "Matcha Latte":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "145")
    elif item_prod.get() == "Carbonara":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "180")
    elif item_prod.get() == "French Toast":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "130")
    elif item_prod.get() == "Caramel Macchiato":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "135")
    elif item_prod.get() == "Tiramisu":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "180")
    elif item_prod.get() == "Strawberry Cheesecake":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "200")
    elif item_prod.get() == "Blueberry Cheesecake":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "200")
    elif item_prod.get() == "Caesar Salad":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "250")
    elif item_prod.get() == "Fries":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "100")
    elif item_prod.get() == "Chicken Alfredo":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "220")
    elif item_prod.get() == "Mushroom Soup":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "150")
    elif item_prod.get() == "Grilled Cheese Sandwich":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "120")
    elif item_prod.get() == "Iced Americano":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "100")
    elif item_prod.get() == "Chocolate Cake":
        price_entry.delete(0, tk.END)
        price_entry.insert(0, "150")
    
    return


window = tk.Tk()
window.title("Cafe Ordering System")
window.geometry("1225x500")
window.configure(bg="pink")

# Form Title
title = tk.Label(window, text="Velvet Beans Cafe", font=("Century Gothic", 25, "bold"), bg="pink")
title.grid(row=0, column=0, columnspan=6, pady=10)

# Frame
genframe = tk.Frame(window, bg="pink", bd=3, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)

# Customer Name Entry
cname_entry = tk.Entry(genframe, font=("Poppins", 15))
cname_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = tk.Label(genframe, text="Customer Name", font=("Poppins", 13, "italic"), bg="pink")
cname_label.grid(row=3, column=1, columnspan=2)

# Product Combobox
product_label = tk.Label(genframe, text="Product", font=("Poppins", 13, "italic"), bg="pink")
product_label.grid(row=3, column=3, columnspan=2)

item_prod = ttk.Combobox(genframe, font=("Poppins", 15))

item_prod['values'] = (
                        "Spanish Latte", 
                       "Matcha Latte",
                       "Caramel Macchiato",
                       "Carbonara", 
                       "French Toast", 
                       "Chocolate Cake",
                       "Tiramisu", 
                       "Strawberry Cheesecake", 
                       "Blueberry Cheesecake", 
                       "Caesar Salad",
                       "Fries",
                       "Chicken Alfredo",
                       "Mushroom Soup",
                       "Grilled Cheese Sandwich",
                       "Iced Americano"
                       )

item_prod.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

# Quantity Entry
qty_entry = tk.Entry(genframe, font=("Poppins", 15))
qty_entry.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

qty_label = tk.Label(genframe, text="Quantity", font=("Poppins", 13, "italic"), bg="pink")
qty_label.grid(row=5, column=1, columnspan=2)

# Price Entry
price_entry = tk.Entry(genframe, font=("Poppins", 15))
price_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

price_label = tk.Label(genframe, text="Price", font=("Poppins", 13, "italic"), bg="pink")
price_label.grid(row=5, column=3, columnspan=2)

# Buttons
submit_btn = tk.Button(window, text="Submit", font=("Poppins", 13, "bold"), bg="lightblue", command=append_excel)
submit_btn.grid(row=6, column=1, padx=(90, 0), pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Poppins", 13, "bold"), bg="lightgreen", command=update)
update_btn.grid(row=6, column=2, padx=(140, 0), pady=(10, 20))

delete_btn = tk.Button(window, text="Delete", bg="red", fg="white",font=("Poppins", 13, "bold"), command=delete)
delete_btn.grid(row=6, column=3, padx=(140, 0), pady=(10, 20))

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"),
    show="headings"
)

for headings in ("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", auto_populate)

item_prod.bind("<<ComboboxSelected>>", auto_price)

display_order()

window.mainloop()