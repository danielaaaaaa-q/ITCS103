import tkinter as tk
import openpyxl as op
from tkinter import messagebox,ttk

def input_validation():
    first = fname_entry.get()
    last = lname_entry.get()
    by = birth_entry.get()

    # required fields
    if not first or not last or not by:
        messagebox.showerror("Error", "Please fill in all required fields.")
        return False

    # must be numeric
    if not by.isdigit():
        messagebox.showerror("Error", "Birth year must be a number.")
        return False
    
    return True

def append_excel():
    if not input_validation():
        return
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    by = int(birth_entry.get())
    age = 2026 - by

    wbk = op.load_workbook("excelDB.xlsx")
    sheet = wbk.active

    new_id = sheet.max_row # Get the next ID based on the current number of rows

    sheet.append([new_id, last, first, middle, by, age])
    wbk.save("excelDB.xlsx")

    messagebox.showinfo("Success", "Record added successfully!")
    display() 

def display():
    wbk = op.load_workbook("excelDB.xlsx")
    sheet = wbk.active

    # clear treeview
    for row in tree.get_children():
        tree.delete(row)

    # insert excel data into treeview
    for row in sheet.iter_rows(values_only=True):
        tree.insert("", "end", values=row)

def select_record(event):
    selected = tree.focus()
    values = tree.item(selected, "values")

    if values:
        lname_entry.delete(0, tk.END)
        fname_entry.delete(0, tk.END)
        mname_entry.delete(0, tk.END)
        birth_entry.delete(0, tk.END)

        lname_entry.insert(0, values[1])
        fname_entry.insert(0, values[2])
        mname_entry.insert(0, values[3])
        birth_entry.insert(0, values[4])

def update_data():
    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error", "Please select a record first.")
        return
    
    if not input_validation():
        return
    
    values = tree.item(selected, "values")
    record_id = values[0]

    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    by = int(birth_entry.get())
    age = 2026 - by

    wbk = op.load_workbook("excelDB.xlsx")
    sheet = wbk.active

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == record_id:
            row[1].value = last
            row[2].value = first
            row[3].value = middle
            row[4].value = by
            row[5].value = age
            
    wbk.save("excelDB.xlsx")

    messagebox.showinfo("Success", "Record updated successfully!")
    display()

def delete_data():
    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error", "Please select a record first.")
        return
    
    values = tree.item(selected, "values")
    record_id = values[0]

    confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this record?")
    if not confirm:
        return

    wbk = op.load_workbook("excelDB.xlsx")
    sheet = wbk.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row[0].value == record_id:
            sheet.delete_rows(i)
            break
            
    wbk.save("excelDB.xlsx")

    messagebox.showinfo("Success", "Record deleted successfully!")
    display()

window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=1,columnspan=2)

update_btn = tk.Button(window, text="Update", command = update_data)
update_btn.grid(row=6, column=2)

button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink", command=append_excel)
button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white", command = delete_data)
delete_btn.grid(row=6, column=3)

tree = ttk.Treeview(window, columns=("ID","Last","First","Middle","BirthYear","Age"), show="headings")
for col in ("ID","Last","First","Middle","BirthYear","Age"):
    tree.heading(col, text=col)
tree.grid(row=8, column=0, columnspan=4)

tree.bind("<<TreeviewSelect>>", select_record)

display()

window.mainloop()
