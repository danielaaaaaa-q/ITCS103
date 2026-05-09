import openpyxl as op
import os

workbook = op.Workbook()
sheet = workbook.active

sheet["A1"] = "ID"
sheet["B1"] = "First Name"
sheet["C1"] = "Last Name"
sheet["D1"] = "Birth Year"
sheet["E1"] = "Age"
sheet["A2"] = 1
sheet["A3"] = 2
sheet["A4"] = 3

workbook.save("favorite_people.xlsx")

print("\nFavorite Person 1")
fname1 = input("Enter first name: ")
lname1 = input("Enter last name: ")
by1 = int(input("Enter birth year: "))

print("\nFavorite Person 2")
fname2 = input("Enter first name: ")
lname2 = input("Enter last name: ")
by2 = int(input("Enter birth year: "))

print("\nFavorite Person 3")
fname3 = input("Enter first name: ")
lname3 = input("Enter last name: ")
by3 = int(input("Enter birth year: "))

print("\nFavorite people recorded successfully!")

#calculating age
age1 = 2026 - by1
age2 = 2026 - by2   
age3 = 2026 - by3

wbk = op.load_workbook("favorite_people.xlsx")
sheet = wbk.active

print("\n=== FAVORITE PEOPLE ===\n")

sheet["B2"] = fname1
sheet["C2"] = lname1
sheet["D2"] = by1
sheet["E2"] = age1

sheet["B3"] = fname2
sheet["C3"] = lname2
sheet["D3"] = by2
sheet["E3"] = age2

sheet["B4"] = fname3
sheet["C4"] = lname3
sheet["D4"] = by3
sheet["E4"] = age3

wbk.save("favorite_people.xlsx")

for rows in sheet.iter_rows(values_only=True):
    print(rows)

print("\n")
os.system("pause")